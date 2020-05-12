
import os
import shutil
import zipfile
import re
import functools
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, colors
from tools import Tools
from log import Log

class OfficeTools():

    def __init__(self, cacheDir):
        super().__init__()
        self.cacheDir = cacheDir

    def extractPics(self, docxFilePath):

        cacheRoot = self.cacheDir
        name = Tools.getFileName(docxFilePath)
        cacheDir = self.cacheDir

        tmpFilePath = Tools.joinPath(cacheDir, '{}.docx'.format(name))
        shutil.copyfile(docxFilePath, tmpFilePath)
        zipFilePath = Tools.joinPath(cacheDir, '{}.zip'.format(name))
        if os.path.exists(zipFilePath):
            os.remove(zipFilePath)
        os.rename(tmpFilePath, zipFilePath)

        extractDir = Tools.joinPath(cacheDir, 'extract')
        Log.i('extracting {} to {}'.format(docxFilePath, extractDir))
        zipF = zipfile.ZipFile(zipFilePath, 'r')
        for file in zipF.namelist():
            zipF.extract(file, extractDir)
        zipF.close()

        imgStoreDir = Tools.joinPath(cacheDir, 'images')
        shutil.rmtree(imgStoreDir, ignore_errors=True)
        Tools.createDir(imgStoreDir)
        mediaDir = Tools.joinPath(extractDir, 'word', 'media')
        r = []
        for fileName in os.listdir(mediaDir):
            srcFileName = Tools.joinPath(mediaDir, fileName)
            dstFileName = Tools.joinPath(imgStoreDir, fileName)
            shutil.copyfile(srcFileName, dstFileName)
            os.remove(srcFileName)
            r.append(dstFileName)

        reg = re.compile(r'image(\d+)\.')
        def cmp(e1, e2):
            index1 = reg.search(e1).group(1)
            index2 = reg.search(e2).group(1)
            return int(index1) - int(index2)
        r.sort(key = functools.cmp_to_key(cmp))

        Tools.removeAllFiles(extractDir)

        return r, imgStoreDir

    @classmethod
    def writeIntoExcel(cls, r, excelFilePath):

        Log.i('writing into excel: ' + excelFilePath)

        book = Workbook()
        sheet = book.active
        
        # sheet.column_dimensions['E'].width = 32.0

        border = Border(left = Side(style = 'thin', color = colors.BLACK),
                        right = Side(style = 'thin', color = colors.BLACK),
                        top = Side(style = 'thin', color = colors.BLACK),
                        bottom = Side(style = 'thin', color =  colors.BLACK))

        leftAlign = Alignment(horizontal = 'left', vertical = 'center')
        centerAlign = Alignment(horizontal = 'center', vertical = 'center')

        # headers = ('户编号', '户主（与户主关系）', '', '性别', '身份证号', '人持股数', '合计人数', '合计股数', '备注')
        # sheet.merge_cells('B3:C3')
        # for i in range(len(headers)):
        #     cell = sheet.cell(row = 3, column = i + 1)
        #     cell.border = border
        #     if i == 2:
        #         continue
        #     cell.value = headers[i]
        #     cell.alignment = centerAlign

        def getCellAlign(cidx):
            return centerAlign
        
        startRow = 4
        for i in range(len(r)):
            rr = r[i]
            for j in range(len(rr)):
                txt = rr[j]
                ridx = startRow + i
                cidx = j + 1
                cell = sheet.cell(row = ridx, column = cidx)
                cell.value = txt
                cell.alignment = getCellAlign(j)
                cell.border = border
        book.save(excelFilePath)

